"""Import real data from Hike POS xlsx exports.

Reads from devtools/data/real/ by default.

Usage:
    # Import vendors first, then products
    PYTHONPATH=backend:. uv run python -m devtools.scripts.import_hike --vendors
    PYTHONPATH=backend:. uv run python -m devtools.scripts.import_hike --products

    # Or both at once (vendors imported first automatically)
    PYTHONPATH=backend:. uv run python -m devtools.scripts.import_hike --vendors --products
"""

import argparse
import asyncio
import re
from datetime import UTC, datetime
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "real"


def _strip_html(text: str | None) -> str:
    if not text:
        return ""
    clean = re.sub(r"<[^>]+>", "", str(text)).strip()
    return "" if clean == "-" else clean


async def import_vendors(file_path: Path | None = None) -> dict[str, str]:
    """Import vendors from Supplier.xlsx. Returns {company_name: vendor_id}."""
    import openpyxl
    from catalog.infrastructure.vendor_repo import vendor_repo

    from catalog.domain.vendor import Vendor
    from shared.infrastructure.logging_config import org_id_var

    path = file_path or DATA_DIR / "Supplier.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    now = datetime.now(UTC).isoformat()

    name_to_id: dict[str, str] = {}
    created = 0
    skipped = 0

    for row in rows[1:]:
        data = dict(zip(headers, row, strict=False))
        company = (data.get("Company name") or "").strip()
        if not company:
            continue

        existing = await vendor_repo.find_by_name(company)
        if existing:
            name_to_id[company] = existing.id
            skipped += 1
            continue

        vendor = Vendor(
            name=company,
            contact_name=f"{data.get('First name', '') or ''} {data.get('Last name', '') or ''}".strip(),
            email=(data.get("Email") or "").strip(),
            phone=(data.get("Phone") or data.get("Mobile") or "").strip(),
            address=", ".join(
                filter(
                    None,
                    [
                        (data.get("Physical address") or "").strip(),
                        (data.get("Physical city") or "").strip(),
                        (data.get("Physical state / region") or "").strip(),
                        (data.get("Physical post / zip code") or "").strip(),
                    ],
                )
            ),
            organization_id=org_id_var.get(),
            created_at=now,
        )
        await vendor_repo.insert(vendor)
        name_to_id[company] = vendor.id
        created += 1

    print(f"Vendors: {created} created, {skipped} already existed")
    wb.close()
    return name_to_id


async def import_products(
    file_path: Path | None = None,
    vendor_map: dict[str, str] | None = None,
) -> None:
    """Import products from Products 2.xlsx."""
    import openpyxl
    from catalog.infrastructure.department_repo import department_repo
    from catalog.infrastructure.vendor_item_repo import vendor_item_repo
    from catalog.infrastructure.vendor_repo import vendor_repo

    from catalog.application.sku_lifecycle import create_product_with_sku
    from catalog.domain.vendor_item import VendorItem
    from documents.application.import_parser import infer_uom
    from shared.infrastructure.logging_config import org_id_var

    path = file_path or DATA_DIR / "Products 2.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    now = datetime.now(UTC).isoformat()

    # Build department lookup: name → (id, name)
    all_depts = await department_repo.list_all()
    dept_by_name = {d.name: d for d in all_depts}

    # Build vendor lookup if not provided
    if vendor_map is None:
        vendor_map = {}
        all_vendors = await vendor_repo.list_all()
        for v in all_vendors:
            vendor_map[v.name] = v.id

    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows[1:], start=2):
        data = dict(zip(headers, row, strict=False))

        name = (data.get("Name") or "").strip()
        if not name:
            continue

        product_type = (data.get("Product type") or "").strip()
        dept = dept_by_name.get(product_type)
        if not dept:
            errors.append(
                f"Row {i}: unknown department '{product_type}' for '{name}'"
            )
            continue

        # Price/cost
        cost = float(data.get("Pittsburgh Store_Cost price") or 0)
        price_ex_tax = float(
            data.get("Pittsburgh Store_Price Excluding Tax") or 0
        )
        price = (
            price_ex_tax
            if price_ex_tax > 0
            else float(data.get("Pittsburgh Store_Retail price") or 0)
        )

        # Stock
        qty = float(
            data.get("Pittsburgh Store_Stock on hand")
            or data.get("Pittsburgh Store_Stock")
            or 0
        )
        min_stock = int(float(data.get("Pittsburgh Store_Reorder level") or 0))

        # Barcode — skip validation for non-standard barcodes
        barcode_raw = str(data.get("Barcode") or "").strip()
        # If barcode is purely digits and fails check-digit validation, store as vendor_barcode only
        barcode = barcode_raw or None
        if barcode and barcode.isdigit() and len(barcode) not in (12, 13):
            barcode = None  # non-standard length, skip barcode field

        # UOM from product name
        base_unit, sell_uom, pack_qty = infer_uom(name)

        # Description
        description = _strip_html(data.get("Description"))

        try:
            sku = await create_product_with_sku(
                category_id=dept.id,
                category_name=dept.name,
                name=name,
                description=description,
                price=price,
                cost=cost,
                quantity=qty,
                min_stock=min_stock,
                barcode=barcode,
                base_unit=base_unit,
                sell_uom=sell_uom,
                pack_qty=pack_qty,
                user_id="import",
                user_name="Hike Import",
            )
            created += 1

            # Link vendor item if supplier info present
            supplier_name = (data.get("Supplier name") or "").strip()
            vendor_sku = str(data.get("SKU") or "").strip()
            vendor_id = vendor_map.get(supplier_name) if supplier_name else None

            if vendor_id and vendor_sku:
                vi = VendorItem(
                    vendor_id=vendor_id,
                    vendor_name=supplier_name,
                    sku_id=sku.id,
                    vendor_sku=vendor_sku,
                    cost=cost,
                    is_preferred=True,
                    organization_id=org_id_var.get(),
                    created_at=now,
                    updated_at=now,
                )
                await vendor_item_repo.insert(vi)

            if created % 100 == 0:
                print(f"  ... {created} products imported")

        except Exception as e:
            err_msg = f"Row {i}: '{name}' — {e}"
            errors.append(err_msg)
            skipped += 1

    print(f"Products: {created} created, {skipped} skipped")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for err in errors[:20]:
            print(f"  {err}")
        if len(errors) > 20:
            print(f"  ... and {len(errors) - 20} more")

    wb.close()


async def main(do_vendors: bool, do_products: bool) -> None:
    from devtools.scripts.company import ORG
    from shared.infrastructure.db import close_db, init_db
    from shared.infrastructure.logging_config import org_id_var, user_id_var

    await init_db()

    # Set org context so repos scope correctly
    org_id_var.set(ORG.id)
    user_id_var.set("import")

    try:
        vendor_map = None
        if do_vendors:
            vendor_map = await import_vendors()

        if do_products:
            await import_products(vendor_map=vendor_map)
            from shared.infrastructure.db.base import get_database_manager

            await (
                get_database_manager().catalog.recompute_department_sku_counts(
                    ORG.id
                )
            )
            print("Recomputed department SKU counts.")

        print("\nDone.")
    finally:
        await close_db()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Hike POS data")
    parser.add_argument(
        "--vendors",
        action="store_true",
        help="Import vendors from Supplier.xlsx",
    )
    parser.add_argument(
        "--products",
        action="store_true",
        help="Import products from Products 2.xlsx",
    )
    args = parser.parse_args()

    if not args.vendors and not args.products:
        parser.error("Specify --vendors, --products, or both")

    asyncio.run(main(args.vendors, args.products))
